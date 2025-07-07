import streamlit as st
import pandas as pd
import re
import io
import zipfile
import os
from datetime import datetime
import unicodedata
import pymupdf as fitz
import pikepdf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# Configuração da página
st.set_page_config(
    page_title="Extrator de Faturas PDF",
    page_icon="⚡",
    layout="wide"
)

# CSS personalizado
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f4e79;
        text-align: center;
        margin-bottom: 2rem;
        font-weight: bold;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
    }
    .info-box {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 15px;
        margin: 1rem 0;
        color: white;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    .feature-card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 4px solid #1f4e79;
        margin: 1rem 0;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
    }
    .metric-card {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        text-align: center;
        margin: 0.5rem;
    }
    .success-box {
        background: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
    .error-box {
        background: #f8d7da;
        border: 1px solid #f5c6cb;
        color: #721c24;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

class ZipProcessor:
    def __init__(self):
        self.pdfs_encontrados = []
        self.pdfs_dist_ee = []
        
    def processar_zip(self, zip_bytes):
        """Processa arquivo ZIP e extrai PDFs que contenham DIST_EE"""
        try:
            pdfs_extraidos = []
            
            with zipfile.ZipFile(io.BytesIO(zip_bytes), 'r') as zip_ref:
                # Listar todos os arquivos no ZIP
                todos_arquivos = zip_ref.namelist()
                self.pdfs_encontrados = [f for f in todos_arquivos if f.lower().endswith('.pdf')]
                
                # Filtrar apenas PDFs com DIST_EE no nome
                pdfs_dist_ee = [f for f in self.pdfs_encontrados if 'DIST_EE' in f.upper()]
                self.pdfs_dist_ee = pdfs_dist_ee
                
                # Extrair conteúdo dos PDFs filtrados
                for nome_arquivo in pdfs_dist_ee:
                    try:
                        # Extrair arquivo do ZIP
                        pdf_bytes = zip_ref.read(nome_arquivo)
                        
                        # Criar objeto tipo arquivo para compatibilidade
                        pdf_info = {
                            'name': os.path.basename(nome_arquivo),
                            'bytes': pdf_bytes,
                            'size': len(pdf_bytes)
                        }
                        
                        pdfs_extraidos.append(pdf_info)
                        
                    except Exception as e:
                        st.warning(f"Erro ao extrair {nome_arquivo}: {str(e)}")
                        continue
            
            return pdfs_extraidos, len(self.pdfs_encontrados), len(self.pdfs_dist_ee)
            
        except Exception as e:
            st.error(f"Erro ao processar arquivo ZIP: {str(e)}")
            return [], 0, 0
    
    def get_estatisticas(self):
        """Retorna estatísticas do processamento do ZIP"""
        return {
            'total_pdfs': len(self.pdfs_encontrados),
            'pdfs_dist_ee': len(self.pdfs_dist_ee),
            'pdfs_outros': len(self.pdfs_encontrados) - len(self.pdfs_dist_ee),
            'lista_dist_ee': self.pdfs_dist_ee,
            'lista_outros': [f for f in self.pdfs_encontrados if f not in self.pdfs_dist_ee]
        }
    def __init__(self):
        self.texto_completo = ""
        self.info_extraida = {}
        
    def extrair_texto_pdf(self, pdf_bytes):
        """Extrai texto do PDF usando PyMuPDF e fallback para PDFs protegidos"""
        try:
            # Tentativa com PyMuPDF
            with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
                texto = ""
                for pagina in doc:
                    texto += pagina.get_text()
                return texto
        except:
            try:
                # Fallback para PDFs protegidos
                senhas_comuns = ["", "33042", "56993", "60869", "08902", "3304", "5699", "6086", "0890"]
                
                for senha in senhas_comuns:
                    try:
                        with pikepdf.open(io.BytesIO(pdf_bytes), password=senha) as pdf:
                            with fitz.open(stream=pdf.save(), filetype="pdf") as doc:
                                texto = ""
                                for pagina in doc:
                                    texto += pagina.get_text()
                                return texto
                    except:
                        continue
                        
                return "ERRO: Não foi possível extrair texto do PDF"
            except Exception as e:
                return f"ERRO: {str(e)}"
    
    def normalizar_texto(self, texto):
        """Normaliza texto removendo acentos e caracteres especiais"""
        if not isinstance(texto, str):
            return ""
        texto = texto.lower().strip()
        texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
        return texto
    
    def extrair_numero_instalacao(self, texto):
        """Extrai número de instalação usando regex"""
        padroes = [
            r'(?:n[°º]?\s*(?:da\s*)?instala[cç][aã]o|instala[cç][aã]o|n[°º]?\s*instala[cç][aã]o)[:\s]*([0-9]{8,12})',
            r'(?:unid[\.\s]*consumidora|uc|c[oó]digo[\s]*do[\s]*cliente)[:\s]*([0-9]{8,12})',
            r'(?:cliente|consumidor)[:\s]*([0-9]{8,12})',
            r'([0-9]{10,12})(?=\s*(?:tarifa|grupo|modalidade))',
        ]
        
        for padrao in padroes:
            matches = re.finditer(padrao, texto, re.IGNORECASE)
            for match in matches:
                numero = match.group(1)
                if len(numero) >= 8:
                    return numero
        return ""
    
    def extrair_referencia(self, texto):
        """Extrai referência (mês/ano) usando regex"""
        padroes = [
            r'(?:refer[eê]ncia|per[ií]odo|compet[eê]ncia)[:\s]*([0-9]{1,2}[\/\-][0-9]{4})',
            r'(?:refer[eê]ncia|per[ií]odo|compet[eê]ncia)[:\s]*([a-z]{3,9}[\/\-][0-9]{4})',
            r'([0-9]{1,2}[\/\-][0-9]{4})',
            r'(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)[a-z]*[\/\-\s]*([0-9]{4})',
        ]
        
        for padrao in padroes:
            matches = re.finditer(padrao, texto, re.IGNORECASE)
            for match in matches:
                if len(match.groups()) == 1:
                    return match.group(1)
                else:
                    return f"{match.group(1)}/{match.group(2)}"
        return ""
    
    def extrair_secao_itens_fatura(self, texto):
        """Extrai seção que contém a tabela de itens da fatura"""
        # Padrão para início da seção
        padrao_inicio = r'(?i)(itens?\s*d[ae]\s*fatura|detalhes?\s*d[ao]\s*faturamento)'
        
        # Padrão para fim da seção
        padroes_fim = [
            r'(?i)(total\s*(?:da\s*)?fatura|total\s*a\s*pagar|total\s*geral|valor\s*total)',
            r'(?i)(informa[cç][oõ]es?\s*importantes?|observa[cç][oõ]es?)',
            r'(?i)(hist[oó]rico\s*de\s*consumo|dados?\s*t[eé]cnicos?)',
        ]
        
        # Encontrar início
        match_inicio = re.search(padrao_inicio, texto)
        if not match_inicio:
            return ""
        
        inicio = match_inicio.start()
        trecho = texto[inicio:]
        
        # Encontrar fim
        fim = len(trecho)
        for padrao_fim in padroes_fim:
            match_fim = re.search(padrao_fim, trecho)
            if match_fim:
                fim = min(fim, match_fim.start() + 200)  # Incluir um pouco após o total
        
        return trecho[:fim]
    
    def limpar_valor_numerico(self, valor_str):
        """Converte string de valor para float tratando formatos brasileiros"""
        if not isinstance(valor_str, str):
            return 0.0
        
        # Remove espaços e caracteres não numéricos exceto vírgula, ponto e sinal negativo
        valor_limpo = re.sub(r'[^\d,.\-]', '', valor_str.strip())
        
        # Trata valores negativos com sinal no final
        if valor_limpo.endswith('-'):
            valor_limpo = '-' + valor_limpo[:-1]
        
        # Converte formato brasileiro (1.234,56) para float
        if ',' in valor_limpo and '.' in valor_limpo:
            # Formato: 1.234,56
            valor_limpo = valor_limpo.replace('.', '').replace(',', '.')
        elif ',' in valor_limpo:
            # Pode ser 1234,56 ou 1,234 - verificar contexto
            partes = valor_limpo.split(',')
            if len(partes[-1]) == 2:  # Provavelmente centavos
                valor_limpo = valor_limpo.replace(',', '.')
            else:  # Provavelmente milhares
                valor_limpo = valor_limpo.replace(',', '')
        
        try:
            return float(valor_limpo)
        except:
            return 0.0
    
    def extrair_tabela_itens(self, secao_itens):
        """Extrai tabela de itens da seção usando regex e heurísticas"""
        if not secao_itens:
            return pd.DataFrame()
        
        linhas = secao_itens.split('\n')
        dados_tabela = []
        
        # Padrões para identificar linhas de dados
        padrao_linha_item = r'([A-Za-zÀ-ÿ\s\./\-]+)\s+([A-Za-z]{1,6})\s+([\d\.,\-]+)\s+([\d\.,\-]+)\s+([\d\.,\-]+)'
        padrao_valor_unico = r'([A-Za-zÀ-ÿ\s\./\-]+)\s+([\d\.,\-]+)$'
        
        for linha in linhas:
            linha = linha.strip()
            if not linha or len(linha) < 10:
                continue
            
            # Ignorar cabeçalhos e linhas de separação
            if any(palavra in linha.lower() for palavra in ['item', 'unidade', 'quantidade', 'valor', 'unitário', '---', '===']):
                continue
            
            # Tentar extrair linha completa com múltiplas colunas
            match_completo = re.search(padrao_linha_item, linha)
            if match_completo:
                item = match_completo.group(1).strip()
                unidade = match_completo.group(2).strip()
                quantidade = self.limpar_valor_numerico(match_completo.group(3))
                valor_unitario = self.limpar_valor_numerico(match_completo.group(4))
                valor_total = self.limpar_valor_numerico(match_completo.group(5))
                
                dados_tabela.append({
                    'Item': item,
                    'Unidade': unidade,
                    'Quantidade': quantidade,
                    'Valor_Unitario': valor_unitario,
                    'Valor_Total': valor_total,
                    'PIS_COFINS': '',
                    'Base_ICMS': '',
                    'Aliquota_ICMS': '',
                    'ICMS': '',
                    'Tarifa_Unitaria': ''
                })
            else:
                # Tentar extrair linha com apenas item e valor
                match_simples = re.search(padrao_valor_unico, linha)
                if match_simples:
                    item = match_simples.group(1).strip()
                    valor_total = self.limpar_valor_numerico(match_simples.group(2))
                    
                    # Filtrar linhas irrelevantes
                    if len(item) > 3 and not any(palavra in item.lower() for palavra in ['total', 'subtotal', 'página', 'folha']):
                        dados_tabela.append({
                            'Item': item,
                            'Unidade': '',
                            'Quantidade': '',
                            'Valor_Unitario': '',
                            'Valor_Total': valor_total,
                            'PIS_COFINS': '',
                            'Base_ICMS': '',
                            'Aliquota_ICMS': '',
                            'ICMS': '',
                            'Tarifa_Unitaria': ''
                        })
        
        if dados_tabela:
            df = pd.DataFrame(dados_tabela)
            # Remover linhas duplicadas e ordenar por valor
            df = df.drop_duplicates(subset=['Item']).reset_index(drop=True)
            return df
        
        return pd.DataFrame()
    
    def processar_pdf(self, pdf_bytes, nome_arquivo):
        """Processa um PDF e extrai todas as informações"""
        resultado = {
            'nome_arquivo': nome_arquivo,
            'numero_instalacao': '',
            'referencia': '',
            'tabela_itens': pd.DataFrame(),
            'erro': ''
        }
        
        try:
            # Extrair texto do PDF
            texto = self.extrair_texto_pdf(pdf_bytes)
            
            if texto.startswith('ERRO:'):
                resultado['erro'] = texto
                return resultado
            
            # Extrair informações básicas
            resultado['numero_instalacao'] = self.extrair_numero_instalacao(texto)
            resultado['referencia'] = self.extrair_referencia(texto)
            
            # Extrair seção de itens
            secao_itens = self.extrair_secao_itens_fatura(texto)
            
            # Extrair tabela
            resultado['tabela_itens'] = self.extrair_tabela_itens(secao_itens)
            
            if resultado['tabela_itens'].empty:
                resultado['erro'] = 'Tabela de itens não encontrada ou vazia'
            
        except Exception as e:
            resultado['erro'] = f'Erro ao processar PDF: {str(e)}'
        
        return resultado

class ExcelGenerator:

class PDFExtractor:
    def __init__(self):
        self.wb = None
        
    def criar_workbook(self):
        """Cria um novo workbook Excel"""
        self.wb = openpyxl.Workbook()
        # Remove a planilha padrão
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
    
    def criar_aba_resumo(self, dados_resumo):
        """Cria aba de resumo com informações consolidadas"""
        ws = self.wb.create_sheet(title="Resumo")
        
        # Cabeçalhos
        cabecalhos = ['Nome_Arquivo', 'Numero_Instalacao', 'Referencia', 'Total_Itens', 'Valor_Total', 'Status']
        
        # Estilo para cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Inserir cabeçalhos
        for col, header in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Inserir dados
        for row, dados in enumerate(dados_resumo, 2):
            ws.cell(row=row, column=1, value=dados['nome_arquivo'])
            ws.cell(row=row, column=2, value=dados['numero_instalacao'])
            ws.cell(row=row, column=3, value=dados['referencia'])
            ws.cell(row=row, column=4, value=len(dados.get('tabela_itens', [])))
            
            # Calcular valor total
            valor_total = 0
            if not dados.get('tabela_itens', pd.DataFrame()).empty:
                valor_total = dados['tabela_itens']['Valor_Total'].sum()
            ws.cell(row=row, column=5, value=valor_total)
            
            # Status
            status = "Sucesso" if dados.get('erro', '') == '' else "Erro"
            ws.cell(row=row, column=6, value=status)
            
            # Colorir linha baseado no status
            if status == "Erro":
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="ffcccc", end_color="ffcccc", fill_type="solid")
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def criar_aba_pdf(self, dados_pdf):
        """Cria aba individual para cada PDF"""
        nome_aba = dados_pdf['nome_arquivo'][:31]  # Limite do Excel para nome de aba
        
        ws = self.wb.create_sheet(title=nome_aba)
        
        # Informações do cabeçalho
        ws.cell(row=1, column=1, value="Arquivo:").font = Font(bold=True)
        ws.cell(row=1, column=2, value=dados_pdf['nome_arquivo'])
        
        ws.cell(row=2, column=1, value="Número Instalação:").font = Font(bold=True)
        ws.cell(row=2, column=2, value=dados_pdf['numero_instalacao'])
        
        ws.cell(row=3, column=1, value="Referência:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=dados_pdf['referencia'])
        
        if dados_pdf.get('erro'):
            ws.cell(row=4, column=1, value="Erro:").font = Font(bold=True, color="FF0000")
            ws.cell(row=4, column=2, value=dados_pdf['erro'])
            return
        
        # Tabela de itens
        if not dados_pdf['tabela_itens'].empty:
            df = dados_pdf['tabela_itens']
            
            # Cabeçalhos da tabela (linha 6)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
            
            for col, header in enumerate(df.columns, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Dados da tabela
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx + 7, column=col_idx, value=value)
                    
                    # Formatação especial para valores numéricos
                    if col_idx in [3, 4, 5]:  # Colunas numéricas
                        if isinstance(value, (int, float)) and value != 0:
                            cell.number_format = '#,##0.00'
            
            # Bordas para a tabela
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(6, len(df) + 7):
                for col in range(1, len(df.columns) + 1):
                    ws.cell(row=row, column=col).border = thin_border
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def salvar_workbook(self):
        """Salva workbook em BytesIO"""
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer

# Interface Streamlit
def main():
    # Cabeçalho
    st.markdown('<h1 class="main-header">⚡ Extrator de Faturas PDF - DIST_EE</h1>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Extraia automaticamente informações de faturas DIST_EE de arquivos ZIP ou PDFs individuais</p>', unsafe_allow_html=True)
    
    # Informações sobre funcionalidades
    with st.expander("ℹ️ Como usar esta ferramenta", expanded=False):
        st.markdown("""
        <div class="feature-card">
        <h4>🎯 O que esta ferramenta faz:</h4>
        <ul>
            <li><strong>Aceita arquivos ZIP</strong> e extrai apenas PDFs com "DIST_EE" no nome</li>
            <li><strong>Processa PDFs individuais</strong> que contenham "DIST_EE"</li>
            <li><strong>Extrai número de instalação</strong> automaticamente</li>
            <li><strong>Identifica a referência</strong> (mês/ano) da fatura</li>
            <li><strong>Localiza e extrai a tabela "Itens de Fatura"</strong></li>
            <li><strong>Gera planilha Excel formatada</strong> com os dados</li>
            <li><strong>Funciona offline</strong> - sem necessidade de API externa</li>
        </ul>
        </div>
        
        <div class="feature-card">
        <h4>📁 Tipos de arquivo aceitos:</h4>
        <ul>
            <li><strong>Arquivos ZIP:</strong> O sistema irá extrair e processar apenas os PDFs que contenham "DIST_EE" no nome</li>
            <li><strong>PDFs individuais:</strong> Serão processados apenas se o nome contiver "DIST_EE"</li>
        </ul>
        </div>
        """, unsafe_allow_html=True)
    
    # Estatísticas da sessão
    if 'pdfs_processados' not in st.session_state:
        st.session_state.pdfs_processados = 0
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <h3>📄 {st.session_state.pdfs_processados}</h3>
            <p>PDFs processados nesta sessão</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <h3>🆓 100%</h3>
            <p>Gratuito e offline</p>
        </div>
        """, unsafe_allow_html=True)
    
    # Upload de arquivos
    st.markdown("### 📂 Selecione seus arquivos")
    
    # Abas para diferentes tipos de upload
    tab1, tab2 = st.tabs(["📦 Arquivo ZIP", "📄 PDFs Individuais"])
    
    uploaded_files = []
    zip_stats = None
    
    with tab1:
        st.markdown("**Envie um arquivo ZIP contendo PDFs de faturas**")
        uploaded_zip = st.file_uploader(
            "Selecione arquivo ZIP",
            type=['zip'],
            help="O sistema extrairá apenas PDFs que contenham 'DIST_EE' no nome"
        )
        
        if uploaded_zip:
            # Processar ZIP
            zip_processor = ZipProcessor()
            zip_bytes = uploaded_zip.read()
            pdfs_extraidos, total_pdfs, pdfs_dist_ee = zip_processor.processar_zip(zip_bytes)
            zip_stats = zip_processor.get_estatisticas()
            
            if pdfs_extraidos:
                uploaded_files = pdfs_extraidos
                
                # Mostrar estatísticas do ZIP
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("📄 Total PDFs no ZIP", total_pdfs)
                with col2:
                    st.metric("⚡ PDFs DIST_EE", pdfs_dist_ee)
                with col3:
                    st.metric("🚫 PDFs ignorados", total_pdfs - pdfs_dist_ee)
                
                # Mostrar detalhes dos arquivos encontrados
                if zip_stats['pdfs_dist_ee'] > 0:
                    with st.expander(f"✅ {zip_stats['pdfs_dist_ee']} PDFs DIST_EE encontrados"):
                        for pdf in zip_stats['lista_dist_ee']:
                            st.write(f"📄 {os.path.basename(pdf)}")
                
                if zip_stats['pdfs_outros'] > 0:
                    with st.expander(f"🚫 {zip_stats['pdfs_outros']} PDFs ignorados (sem DIST_EE)"):
                        for pdf in zip_stats['lista_outros']:
                            st.write(f"📄 {os.path.basename(pdf)}")
            else:
                st.warning("❌ Nenhum PDF com 'DIST_EE' encontrado no arquivo ZIP")
    
    with tab2:
        st.markdown("**Envie PDFs individuais (apenas arquivos com 'DIST_EE' no nome serão processados)**")
        uploaded_pdfs = st.file_uploader(
            "Selecione PDFs individuais",
            type=['pdf'],
            accept_multiple_files=True,
            help="Apenas PDFs que contenham 'DIST_EE' no nome serão processados"
        )
        
        if uploaded_pdfs:
            # Filtrar apenas PDFs com DIST_EE
            pdfs_dist_ee = [pdf for pdf in uploaded_pdfs if 'DIST_EE' in pdf.name.upper()]
            pdfs_ignorados = [pdf for pdf in uploaded_pdfs if 'DIST_EE' not in pdf.name.upper()]
            
            if pdfs_dist_ee:
                # Converter para formato compatível
                uploaded_files = []
                for pdf in pdfs_dist_ee:
                    uploaded_files.append({
                        'name': pdf.name,
                        'bytes': pdf.read(),
                        'size': pdf.size
                    })
                
                # Mostrar estatísticas
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("✅ PDFs DIST_EE", len(pdfs_dist_ee))
                with col2:
                    st.metric("🚫 PDFs ignorados", len(pdfs_ignorados))
                
                # Mostrar listas
                if pdfs_dist_ee:
                    with st.expander(f"✅ {len(pdfs_dist_ee)} PDFs que serão processados"):
                        for pdf in pdfs_dist_ee:
                            st.write(f"📄 {pdf.name} ({pdf.size:,} bytes)")
                
                if pdfs_ignorados:
                    with st.expander(f"🚫 {len(pdfs_ignorados)} PDFs ignorados (sem DIST_EE)"):
                        for pdf in pdfs_ignorados:
                            st.write(f"📄 {pdf.name}")
            else:
                st.warning("❌ Nenhum PDF com 'DIST_EE' encontrado nos arquivos selecionados")
    
    # Informações sobre arquivos selecionados
    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} arquivo(s) DIST_EE pronto(s) para processamento")
    
    # Botão de processamento
    if st.button("🚀 Processar PDFs", type="primary", disabled=not uploaded_files):
        
        # Inicializar extrator e gerador Excel
        extractor = PDFExtractor()
        excel_gen = ExcelGenerator()
        excel_gen.criar_workbook()
        
        # Lista para armazenar resultados
        resultados = []
        
        # Barra de progresso
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Container para resultados em tempo real
        results_container = st.container()
        
        # Processar cada PDF
        for idx, uploaded_file in enumerate(uploaded_files):
            # Atualizar progresso
            progress = (idx + 1) / len(uploaded_files)
            progress_bar.progress(progress)
            
            # Verificar se é objeto tipo dicionário (do ZIP) ou objeto file do Streamlit
            if isinstance(uploaded_file, dict):
                nome_arquivo = uploaded_file['name']
                pdf_bytes = uploaded_file['bytes']
                tamanho_arquivo = uploaded_file['size']
            else:
                nome_arquivo = uploaded_file.name
                pdf_bytes = uploaded_file.read()
                tamanho_arquivo = uploaded_file.size
            
            status_text.text(f"🔍 Processando {idx + 1}/{len(uploaded_files)}: {nome_arquivo}")
            
            # Processar PDF
            resultado = extractor.processar_pdf(pdf_bytes, nome_arquivo)
            resultados.append(resultado)
            
            # Mostrar resultado em tempo real
            with results_container:
                col1, col2, col3 = st.columns([2, 1, 1])
                
                with col1:
                    st.write(f"📄 {uploaded_file.name}")
                
                with col2:
                    if resultado['erro']:
                        st.error("❌ Erro")
                    else:
                        st.success("✅ Sucesso")
                
                with col3:
                    if not resultado['tabela_itens'].empty:
                        st.info(f"📊 {len(resultado['tabela_itens'])} itens")
                    else:
                        st.warning("📊 0 itens")
        
        # Criar abas no Excel
        status_text.text("📊 Gerando planilha Excel...")
        
        # Aba de resumo
        excel_gen.criar_aba_resumo(resultados)
        
        # Aba para cada PDF
        for resultado in resultados:
            excel_gen.criar_aba_pdf(resultado)
        
        # Finalizar
        progress_bar.progress(1.0)
        status_text.text("✅ Processamento concluído!")
        
        # Atualizar contador
        st.session_state.pdfs_processados += len(uploaded_files)
        
        # Estatísticas finais
        sucessos = len([r for r in resultados if not r['erro']])
        erros = len([r for r in resultados if r['erro']])
        total_itens = sum([len(r['tabela_itens']) for r in resultados])
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("✅ Sucessos", sucessos)
        with col2:
            st.metric("❌ Erros", erros)
        with col3:
            st.metric("📊 Total de Itens", total_itens)
        
        # Gerar arquivo Excel para download
        excel_buffer = excel_gen.salvar_workbook()
        
        # Nome do arquivo com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"faturas_extraidas_{timestamp}.xlsx"
        
        # Botão de download
        st.download_button(
            label="📥 Baixar Planilha Excel",
            data=excel_buffer,
            file_name=nome_arquivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Clique para baixar a planilha com todos os dados extraídos"
        )
        
        # Mostrar resumo dos erros se houver
        if erros > 0:
            with st.expander("⚠️ Detalhes dos erros"):
                for resultado in resultados:
                    if resultado['erro']:
                        st.error(f"**{resultado['nome_arquivo']}**: {resultado['erro']}")
        
        # Mensagem final
        if sucessos > 0:
            st.markdown("""
            <div class="success-box">
            <strong>🎉 Processamento concluído com sucesso!</strong><br>
            Sua planilha Excel foi gerada com todas as informações extraídas.
            Cada PDF processado com sucesso tem sua própria aba na planilha.
            </div>
            """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
